"""Build the Drive package master Excel workbook.

One sheet per chapter listing every series with key metadata + validation
status, plus a TOC and a global summary.

Usage:
    python _build_master_workbook.py <out_xlsx_path>
"""
from __future__ import annotations

import json
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Parameterized 2026-05-19 per Decision 0006: derive paths from script location
# (scripts/ -> replicator/ -> Technical/) with RSCD_TECHNICAL env override.
_SCRIPT_DIR = Path(__file__).resolve().parent
_TECHNICAL_DEFAULT = _SCRIPT_DIR.parents[1]   # replicator/scripts -> Technical
TECHNICAL = Path(os.environ.get("RSCD_TECHNICAL", str(_TECHNICAL_DEFAULT)))


def _chapter_label(s_meta: dict) -> str:
    sid = s_meta.get("sid", "")
    ch = s_meta.get("chapter")
    if sid.startswith("ES"):
        return "External_Studies"
    if sid.startswith("AS"):
        return "Analytical_Series"
    if ch is None or ch == "":
        return "Other"
    return f"Chapter_{int(ch):02d}"


def main(out_path: str) -> None:
    reg = json.loads((TECHNICAL / "series_registry.json").read_text(encoding="utf-8"))
    vrpt = json.loads((TECHNICAL / "VALIDATION_REPORT.json").read_text(encoding="utf-8"))

    series = reg.get("series", {})
    vseries = vrpt.get("series", {})

    # Build per-chapter buckets
    buckets: dict[str, list[dict]] = {}
    for sid, meta in series.items():
        meta2 = dict(meta)
        meta2["sid"] = sid
        # attach validation
        v = vseries.get(sid, {})
        meta2["v_status"] = v.get("status", "unknown")
        meta2["v_mae"] = v.get("mae")
        meta2["v_max_abs_err"] = v.get("max_abs_err")
        meta2["v_n_compared"] = v.get("n_compared")
        lab = _chapter_label(meta2)
        buckets.setdefault(lab, []).append(meta2)

    wb = Workbook()
    # Drop default sheet; we'll add TOC first
    ws_toc = wb.active
    ws_toc.title = "TOC"
    _write_toc(ws_toc, buckets)

    ws_sum = wb.create_sheet("Summary")
    _write_summary(ws_sum, series, vseries)

    for lab in sorted(buckets.keys()):
        rows = sorted(buckets[lab], key=lambda r: r["sid"])
        ws = wb.create_sheet(lab[:31])  # Excel limit
        _write_chapter_sheet(ws, lab, rows)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path} ({len(buckets)} chapter sheets, {len(series)} series)")


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def _stylize_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22


def _autosize(ws, headers, sample_rows):
    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        max_w = max(len(str(h)), *(len(str(r.get(h, ""))[:80]) for r in sample_rows[:30]))
        ws.column_dimensions[col].width = min(max_w + 2, 50)


def _write_toc(ws, buckets):
    ws["A1"] = "RSCD v1.0 — Master Workbook"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "Replication and Extension of Shaikh (2016) Capitalism: Competition, Conflict, Crises"
    ws["A2"].font = Font(italic=True, size=11)
    ws["A4"] = "Sheet"
    ws["B4"] = "Series count"
    ws["C4"] = "Description"
    for c in (4,):
        for col in ("A", "B", "C"):
            ws[f"{col}{c}"].font = HEADER_FONT
            ws[f"{col}{c}"].fill = HEADER_FILL
    r = 5
    for lab in sorted(buckets.keys()):
        ws.cell(row=r, column=1, value=lab[:31])
        ws.cell(row=r, column=2, value=len(buckets[lab]))
        ws.cell(row=r, column=3, value=_chapter_desc(lab))
        r += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 80


def _chapter_desc(lab: str) -> str:
    d = {
        "Chapter_02": "Real Competition (industrial production, prices, wages)",
        "Chapter_03": "Microeconomics (firm size, profit dispersion, market structure)",
        "Chapter_04": "Production and costs",
        "Chapter_05": "Exchange, money, and uncertainty",
        "Chapter_06": "Capital and profit (rate of profit, organic composition)",
        "Chapter_07": "Competition and inter-industry dynamics",
        "Chapter_08": "Demand, supply, and price determination",
        "Chapter_09": "Aggregate demand and supply",
        "Chapter_10": "Money, credit, and finance",
        "Chapter_11": "Modern macroeconomic theory",
        "Chapter_13": "Inflation",
        "Chapter_14": "Unemployment",
        "Chapter_15": "Growth and accumulation",
        "Chapter_16": "International trade and exchange rates",
        "Chapter_17": "Crisis and the long downturn",
        "External_Studies": "Replications of cited external econometric studies",
        "Analytical_Series": "Framework-derived series (Ch6 GPIM variants, etc.)",
    }
    return d.get(lab, "")


def _write_summary(ws, series, vseries):
    ws["A1"] = "RSCD v1.0 Summary"
    ws["A1"].font = Font(bold=True, size=14)
    rows = [
        ("Total series", len(series)),
        ("PASS (face-value match)", sum(1 for s in vseries.values() if s.get("status") == "PASS")),
        ("PASS_THEORETICAL", sum(1 for s in vseries.values() if s.get("status") == "PASS_THEORETICAL")),
        ("PASS_DATA_UNAVAILABLE", sum(1 for s in vseries.values() if s.get("status") == "PASS_DATA_UNAVAILABLE")),
        ("PASS_CROSS_SECTIONAL_UNAVAILABLE",
         sum(1 for s in vseries.values() if s.get("status") == "PASS_CROSS_SECTIONAL_UNAVAILABLE")),
        ("Chopped CSVs produced", 109),
        ("Extension workbooks produced", 109),
        ("Verbatim Shaikh quotes (research/)", 118),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16


def _write_chapter_sheet(ws, lab, rows):
    headers = ["SID", "Name", "Chapter", "Figures",
               "Year range (book)", "Year range (extended)",
               "Content type", "Construction", "Status",
               "Units", "Primary source", "Validation status",
               "MAE", "Max abs err", "n compared",
               "Proxy?", "DPR doc", "EPR doc"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _stylize_header(ws, len(headers))
    for r, row in enumerate(rows, start=2):
        def yr(rng):
            if isinstance(rng, list) and len(rng) == 2:
                return f"{rng[0]}-{rng[1]}"
            return ""
        figs = row.get("figures") or []
        if isinstance(figs, list):
            figs_s = ", ".join(str(f) for f in figs)
        else:
            figs_s = str(figs)
        ws.cell(row=r, column=1, value=row.get("sid"))
        ws.cell(row=r, column=2, value=row.get("name"))
        ws.cell(row=r, column=3, value=row.get("chapter"))
        ws.cell(row=r, column=4, value=figs_s)
        ws.cell(row=r, column=5, value=yr(row.get("year_range_book")))
        ws.cell(row=r, column=6, value=yr(row.get("year_range_extension") or row.get("year_range")))
        ws.cell(row=r, column=7, value=row.get("content_type"))
        ws.cell(row=r, column=8, value=row.get("construction"))
        ws.cell(row=r, column=9, value=row.get("status"))
        ws.cell(row=r, column=10, value=row.get("units"))
        ws.cell(row=r, column=11, value=row.get("primary_source"))
        ws.cell(row=r, column=12, value=row.get("v_status"))
        ws.cell(row=r, column=13, value=row.get("v_mae"))
        ws.cell(row=r, column=14, value=row.get("v_max_abs_err"))
        ws.cell(row=r, column=15, value=row.get("v_n_compared"))
        ws.cell(row=r, column=16, value="yes" if row.get("proxy") else "no")
        ws.cell(row=r, column=17, value=row.get("dpr"))
        ws.cell(row=r, column=18, value=row.get("epr"))
    _autosize(ws, headers, [{h: r.get(h.lower().replace(" ", "_"), "") for h in headers} for r in rows])
    # Manual widths for known long fields
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["K"].width = 26
    ws.column_dimensions["Q"].width = 36
    ws.column_dimensions["R"].width = 36
    ws.freeze_panes = "A2"


if __name__ == "__main__":
    # Parameterized 2026-05-19 per Decision 0006: default out derives from RSCD root.
    _RSCD_ROOT = TECHNICAL.parent  # Technical -> RSCD
    _DEFAULT_OUT = (
        _RSCD_ROOT / "Outputs" / "Drive" / "RSCD_v1.0" / "RSCD_v1.0_Master.xlsx"
    )
    out = sys.argv[1] if len(sys.argv) > 1 else str(_DEFAULT_OUT)
    main(out)
