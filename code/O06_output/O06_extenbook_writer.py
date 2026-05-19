"""O06_extenbook_writer — per-series xlsx with [Data, Methodology, Sources, Validation, Provenance].

Reads:
  Technical/data/processed/{SID}.parquet         — the time series
  Technical/docs/series/{SID}_DPR.md             — methodology narrative
  Technical/docs/series/{SID}_EPR.md             — extension narrative (if present)
  Technical/SUBSOURCE_METADATA.json              — source metadata
  Technical/VALIDATION_REPORT.json               — validation row
  Technical/series_registry.json                 — subseries definitions
  Technical/ANU_LEDGER.json                      — artifact provenance

Writes:
  Technical/extenbooks/{SID}_extenbook.xlsx

CLI:
    python O06_extenbook_writer.py --series S201
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from utils import paths  # noqa: E402
from utils.paths import (  # noqa: E402
    DATA_PROCESSED, EXTENBOOK_DIR, DOCS_SERIES, SUBSOURCE_METADATA, REGISTRY, LEDGER, TECHNICAL,
)


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            max_len = max(max_len, min(len(str(v)), max_width))
        ws.column_dimensions[letter].width = max_len + 2


def _write_data_sheet(wb: Workbook, sid: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Data")
    # Pivot to one column per subseries for readability
    pivot = df.pivot_table(index="year", columns="subseries_id", values="value", aggfunc="first")
    pivot = pivot.reset_index()
    ws.append([sid + " — annual values (long form below pivot)"])
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.append([])
    ws.append(list(pivot.columns))
    _style_header(ws, row=3)
    for row in pivot.itertuples(index=False):
        ws.append(list(row))

    # Inline line chart
    if len(pivot) > 0 and len(pivot.columns) > 1:
        chart = LineChart()
        chart.title = f"{sid} — Industrial Production Index (1958=100)"
        chart.y_axis.title = "Index (1958=100)"
        chart.x_axis.title = "Year"
        data_end_col = len(pivot.columns)
        data_start_row = 3
        data_end_row = data_start_row + len(pivot)
        data_ref = Reference(ws, min_col=2, max_col=data_end_col,
                             min_row=data_start_row, max_row=data_end_row)
        cats_ref = Reference(ws, min_col=1, max_col=1,
                             min_row=data_start_row + 1, max_row=data_end_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 24
        chart.height = 12
        ws.add_chart(chart, f"{get_column_letter(data_end_col + 2)}3")

    # Long form below
    long_start = ws.max_row + 4
    ws.cell(row=long_start, column=1, value="Long form (year, value, subseries_id, source_id, units)").font = TITLE_FONT
    headers = ["year", "value", "subseries_id", "source_id", "units"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=long_start + 1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for i, row in enumerate(df[headers].itertuples(index=False), start=long_start + 2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    _autosize(ws)


def _write_methodology_sheet(wb: Workbook, sid: str) -> None:
    ws = wb.create_sheet("Methodology")
    dpr_path = DOCS_SERIES / f"{sid}_DPR.md"
    epr_path = DOCS_SERIES / f"{sid}_EPR.md"
    ws["A1"] = f"{sid} — Methodology"
    ws["A1"].font = TITLE_FONT
    row = 3
    for label, path in (("DPR", dpr_path), ("EPR", epr_path)):
        ws.cell(row=row, column=1, value=f"=== {label}: {path.name} ===").font = Font(bold=True)
        row += 1
        if path.exists():
            for line in path.read_text(encoding="utf-8").splitlines():
                ws.cell(row=row, column=1, value=line)
                row += 1
        else:
            ws.cell(row=row, column=1, value="(not present)")
            row += 1
        row += 2
    ws.column_dimensions["A"].width = 100


def _write_sources_sheet(wb: Workbook, sid: str, subseries_ids: list[str]) -> None:
    ws = wb.create_sheet("Sources")
    meta = json.loads(SUBSOURCE_METADATA.read_text(encoding="utf-8"))["subsources"]
    reg = json.loads(REGISTRY.read_text(encoding="utf-8"))
    sub_defs = reg["series"][sid].get("subseries", {})

    headers = ["subseries_id", "subsource_id", "name", "agency", "url",
               "native_units", "frequency", "license", "retrieval_method"]
    ws.append(headers)
    _style_header(ws, row=1)
    for sub_id in subseries_ids:
        sub = sub_defs.get(sub_id, {})
        ss_id = sub.get("subsource_id", "")
        ss = meta.get(ss_id, {})
        ws.append([
            sub_id, ss_id, ss.get("full_title", ""), ss.get("agency", ""),
            ss.get("url", "") or "",
            ss.get("native_units", ""), ss.get("frequency", ""),
            ss.get("license", ""), ss.get("retrieval_method", ""),
        ])
    _autosize(ws)


def _write_validation_sheet(wb: Workbook, sid: str) -> None:
    ws = wb.create_sheet("Validation")
    ws["A1"] = f"{sid} — Validation"
    ws["A1"].font = TITLE_FONT
    rpt_path = TECHNICAL / "VALIDATION_REPORT.json"
    if not rpt_path.exists():
        ws["A3"] = "VALIDATION_REPORT.json not present."
        return
    row = json.loads(rpt_path.read_text(encoding="utf-8")).get("series", {}).get(sid, {})
    if not row:
        ws["A3"] = f"No validation row for {sid}."
        return
    ws.append([])
    ws.append(["field", "value"])
    _style_header(ws, row=ws.max_row)
    for k, v in row.items():
        ws.append([k, json.dumps(v) if isinstance(v, (list, dict)) else v])
    _autosize(ws)


def _write_provenance_sheet(wb: Workbook, sid: str) -> None:
    ws = wb.create_sheet("Provenance")
    ws["A1"] = f"{sid} — Provenance"
    ws["A1"].font = TITLE_FONT
    led = json.loads(LEDGER.read_text(encoding="utf-8"))
    entry = led.get("series", {}).get(sid, {})
    ws.append([])
    ws.append(["field", "value"])
    _style_header(ws, row=ws.max_row)
    for k, v in entry.items():
        ws.append([k, json.dumps(v, default=str) if isinstance(v, (list, dict)) else v])
    _autosize(ws)


def write_extenbook(sid: str) -> dict:
    proc = DATA_PROCESSED / f"{sid}.parquet"
    if not proc.exists():
        return {"status": "FAIL", "sid": sid, "error": f"processed missing: {proc}"}
    df = pd.read_parquet(proc)
    if "units" not in df.columns:
        df["units"] = ""

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    _write_data_sheet(wb, sid, df)
    _write_methodology_sheet(wb, sid)
    sub_ids = sorted(df["subseries_id"].unique().tolist())
    _write_sources_sheet(wb, sid, sub_ids)
    _write_validation_sheet(wb, sid)
    _write_provenance_sheet(wb, sid)

    EXTENBOOK_DIR.mkdir(parents=True, exist_ok=True)
    out = EXTENBOOK_DIR / f"{sid}_extenbook.xlsx"
    wb.save(out)
    return {"status": "OK", "sid": sid, "output": str(out),
            "size_bytes": out.stat().st_size,
            "sheets": ["Data", "Methodology", "Sources", "Validation", "Provenance"]}


def run() -> dict:
    results = []
    for p in sorted(DATA_PROCESSED.glob("*.parquet")):
        results.append(write_extenbook(p.stem))
    return {"status": "OK", "series": results}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--series", type=str, default=None)
    parser.add_argument("--all", action="store_true")
    args = parser.parse_args()
    if args.series:
        r = write_extenbook(args.series)
    elif args.all:
        r = run()
    else:
        parser.print_help()
        return 0
    print(json.dumps(r, indent=2, default=str))
    return 0 if r.get("status") == "OK" else 1


if __name__ == "__main__":
    raise SystemExit(main())
